from openpyxl import load_workbook
from common.config_path import test_data_path, cases_sheet_name, case_is_col, case_step_name_col


def read_data(file_name):  # section 配置文件里面的片段名 可以根据你的指定来执行具体的用例
    """从Excel读取数据，有返回值"""
    test_data = []
    case_list = []
    wb = load_workbook(file_name)
    total_sheet = wb[cases_sheet_name]
    for row in range(2, total_sheet.max_row + 1):
        # 循环汇总表每一行数据
        is_execute = total_sheet.cell(row, case_is_col).value
        if is_execute == "y":
            # 判断执行，则获取操作步骤的名称
            case_name = total_sheet.cell(row, case_step_name_col).value
            case_list.append(case_name)
    for j in case_list:
        sheet = wb[j]

        for i in range(2, sheet.max_row + 1):
            row_data = dict()
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Modular'] = sheet.cell(i, 2).value
            row_data['ApiName'] = sheet.cell(i, 3).value
            row_data['path'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value
            row_data['Headers'] = sheet.cell(i, 6).value
            row_data['Params'] = sheet.cell(i, 7).value
            row_data['Raw'] = sheet.cell(i, 8).value
            row_data['FormData'] = sheet.cell(i, 9).value
            row_data['Files'] = sheet.cell(i, 10).value
            row_data['ExpectedResult'] = sheet.cell(i, 11).value
            row_data['Extract'] = sheet.cell(i, 12).value
            row_data['Sql'] = sheet.cell(i, 13).value
            test_data.append(row_data)

    wb.close()
    return test_data


if __name__ == '__main__':
    test_datas = read_data(test_data_path)  # 获取测试数据
    for k in test_datas:
        print(k)
