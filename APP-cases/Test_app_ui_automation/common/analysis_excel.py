# coding=utf-8
import time
import openpyxl
import config_path
import os
from openpyxl.styles import PatternFill


class AnalysisExcels:
    """操作excel用例数据文件"""

    def __init__(self, file_name):
        # 设置文件路径
        self.filepath = file_name
        # 获取测试用例excel工作簿
        self.wk = openpyxl.load_workbook(file_name)

    def get_case_data(self):
        """获取excel工作簿中执行用例的数据"""
        # 获取执行用例的工作表名称
        sheet_names = self.get_case_name()
        values = []

        # 循环每个工作表
        for sheet_name in sheet_names:
            # 获取某个工作表某个区间列的数据
            value = self.get_start_col_end_col_value(sheet_name)
            case_data = dict()
            case_data["case_name"] = sheet_name
            case_data["step_data"] = value
            values.append(case_data)

        return values

    # 获取执行用例的sheet页名称
    def get_case_name(self, sheet_name=config_path.cases_sheet_name):
        # 获取汇总用例的工作表表
        sheet = self.wk[sheet_name]
        case_list = []
        # 根据是否执行取到用例的名称
        for row in range(2, sheet.max_row + 1):
            # 循环汇总表每一行数据
            is_execute = sheet.cell(row, config_path.case_is_col).value
            if is_execute == "y":
                # 判断执行，则获取操作步骤的名称
                case_name = sheet.cell(row, config_path.case_step_name_col).value
                case_list.append(case_name)
        return case_list

    # 获取某个工作表某个区间列的数据
    def get_start_col_end_col_value(self, sheet_name, start=config_path.keyword_col, end=config_path.action_col):
        # 获取工作表
        sheet = self.wk[sheet_name]
        values = []
        # 循环对应工作表中的每一行数据
        for row in range(2, sheet.max_row + 1):
            step_data = []
            for col in range(start, end + 1):
                value = sheet.cell(row, col).value
                if value is not None:
                    step_data.append(value)
            values.append(step_data)
        return values

    def write_step_result(self, sheet_name, row, col, result):
        """写入测试步骤的结果"""
        case_sheet = self.wk[sheet_name]
        case_sheet.cell(row, col).value = result
        # 颜色填充   绿色通过   红色失败
        red_fill = PatternFill("solid", fgColor="00FF0000")
        green_fill = PatternFill("solid", fgColor="0000FF00")
        if result == 'FALSE':
            # 如果失败当前单元格填充为红色，否则为绿色
            case_sheet.cell(row, col).fill = red_fill
        else:
            case_sheet.cell(row, col).fill = green_fill
        self.wk.save(self.filepath)

    # 把执行用例的结果写入汇总表
    def write_case_result(self, case_name, sheet_name=config_path.cases_sheet_name, col=config_path.case_result_col):
        sheet = self.wk[sheet_name]
        # 颜色填充   绿色通过   红色失败
        red_fill = PatternFill("solid", fgColor="00FF0000")
        green_fill = PatternFill("solid", fgColor="0000FF00")
        for row in range(2, sheet.max_row + 1):  # 循环汇总表每一行
            is_execute = sheet.cell(row, config_path.case_is_col).value  # 获取"是否执行"所在列的值
            case_name_result = sheet.cell(row, config_path.case_step_name_col).value  # 获取"操作步骤"所在列的值
            if is_execute == "y" and case_name_result == case_name:  # 在汇总表中找到用例所在列
                step_results = self.get_sheet_col_values(case_name)  # 获取该用例所有测试步骤的结果
                if "FALSE" in step_results:
                    # 写入用户测试结果：FALSE
                    sheet.cell(row, col).fill = red_fill
                    sheet.cell(row, col).value = "FALSE"
                    current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                    sheet.cell(row, col - 1).value = current_time
                    self.wk.save(self.filepath)
                elif "PASS" in step_results:
                    # 写入用户测试结果：PASS
                    sheet.cell(row, col).fill = green_fill
                    sheet.cell(row, col).value = "PASS"
                    current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                    sheet.cell(row, col - 1).value = current_time
                    self.wk.save(self.filepath)
                else:
                    continue

    def get_sheet_col_values(self, sheet_name):
        """获取某个工作表某列的所有数据"""
        values = []
        sheet = self.wk[sheet_name]
        for row in range(2, sheet.max_row + 1):
            step_result = sheet.cell(row, config_path.case_step_result).value
            values.append(step_result)
        return values

    # 写入测试步骤执行的时间
    def write_step_time(self, sheet_name, row, col=config_path.action_time_col):
        sheet = self.wk[sheet_name]
        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        sheet.cell(row, col).value = current_time
        self.wk.save(self.filepath)

    # 清除用例汇总表的执行信息
    def clear_excel_value(self):
        sheet = self.wk[config_path.cases_sheet_name]  # 获取用例汇总工作表
        case_list = []
        white_fill = PatternFill(fill_type=None)  # 表格的填充类型为空
        # 清除用例汇总表6、7行的数据
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, 6).value = None
            sheet.cell(row, 7).fill = white_fill
            sheet.cell(row, 7).value = None
            case_name = sheet.cell(row, config_path.case_step_name_col).value
            case_list.append(case_name)
        self.wk.save(self.filepath)
        # 清除用例表7、8行的数据
        for case_name in case_list:
            sheet = self.wk[case_name]
            # 循环对应工作表中的每一行数据
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, 7).value = None
                sheet.cell(row, 8).fill = white_fill
                sheet.cell(row, 8).value = None

            self.wk.save(self.filepath)


if __name__ == '__main__':
    filename = os.path.join(config_path.test_data_dir, "testcases_login.xlsx")  # 获取测试数据路径
    analysis_excels = AnalysisExcels(filename)  # 创建AnalysisExcels对象，用于操作excel
    analysis_excels.clear_excel_value()  # 清除用例表的执行信息
    case_data_ = analysis_excels.get_case_data()  # 获取测试数据
    print(case_data_)
